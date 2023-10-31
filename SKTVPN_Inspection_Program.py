import os
import re
from openpyxl import load_workbook


gw_name = ["DJ_SKTVPN_GW1N" ,
           "DJ_SKTVPN_GW2N" ,
           "DJ_SKTVPN_GW3N" ,
           "DJ_SKTVPN_GW4N" ,
           "DOJ_SKTVPN_GW1N",
           "DOJ_SKTVPN_GW2N",
           "DOJ_SKTVPN_GW3N",
           "DOJ_SKTVPN_GW4N"]

## search log directory
print("### Wait.. Moving Directory ###")
nowdir = os.getcwd()
logdir = "\log"
nextdir = nowdir+logdir
os.chdir(str(nextdir))

loglist = os.listdir(os.getcwd())
print("### Directory Move Successful ###\n")

## Module1 - change filename

print("### Filename Change ###")
for filenum in range(0,8):
    txtfile = open("%s" %loglist[filenum], "r")
    txtline = txtfile.readlines()
    namesplit = 0
    txtfile.close()

    for search_line in range(len(txtline)):
        host_pattern = r'^(DJ|DOJ)_SKTVPN_GW.N#show version'
        search_result = re.match(host_pattern, txtline[search_line])
        if search_result != None:
            hostname = txtline[search_line]
            namesplit = hostname.split(sep = '#',maxsplit = 1)
        else:
            pass
    
    for namecheck in range(0,8):
        if namesplit[0] == gw_name[namecheck]:
            os.rename("%s" %loglist[filenum],"%s.txt" %namesplit[0])
            print("Filename %s -> %s.txt" %(loglist[filenum],namesplit[0]))
        else:
            pass

loglist = os.listdir(os.getcwd())

for namecheck in range(0,8):
    if loglist[namecheck] == "%s.txt" %gw_name[namecheck]:
        pass
    else:
        print("namecheck error")

print("\n")


# Module2 - Extract Memory Utilization_1 #

print("### Extract Memory Utilization_1 ###")
excelpath = "%s\summary.xlsx" %nowdir

wb = load_workbook(excelpath)
ws = wb["Calculator"]

cal_excelrow1 = 7
cal_excelrow2 = 3

loglist = os.listdir(os.getcwd())

for filenum in range(0,8):
    txtfile = open("%s" %loglist[filenum], "r")
    txtline = txtfile.readlines()
    txtfile.close()
    memorysplit = 0

    for search_line in range(len(txtline)):
        host_pattern = r'^%s#show memory platform' %gw_name[filenum]
        search_result = re.match(host_pattern, txtline[search_line])
        if search_result != None:
            memory = txtline[search_line + 10]
            memorysplit = memory.split(sep = ':',maxsplit = 1)

    if memorysplit[0] == '    Used           ' :
        memoryused = int(memorysplit[1])
        print("%s used = %s" %(loglist[filenum],memoryused))
        if cal_excelrow1 <= 10:
            ws["D%s" %cal_excelrow1] = str(memoryused)
            wb.save(excelpath)
            cal_excelrow1 = cal_excelrow1 + 1
        else:
            ws["D%s" %cal_excelrow2] = str(memoryused)
            wb.save(excelpath)
            cal_excelrow2 = cal_excelrow2 + 1

    else:
        print("error")

print("\n")

# Module3 - Extract Memory Utilization_2 #

print("### Extract Memory Utilization_2 ###")
excelpath = "%s\summary.xlsx" %nowdir

wb = load_workbook(excelpath)
ws = wb["Summary"]

sum_excelrow1 = 2

loglist = os.listdir(os.getcwd())

for filenum in range(0,8):
    txtfile = open("%s" %loglist[filenum], "r")
    txtline = txtfile.readlines()
    txtfile.close()

    for search_line in range(len(txtline)):
        host_pattern = r'^%s#show memory processor statistics' %gw_name[filenum]
        search_result = re.match(host_pattern, txtline[search_line])
        if search_result != None:
            for linenum in range(0,4):
                stat_line = txtline[search_line + linenum]
                ws["B%s" %sum_excelrow1] = str(stat_line)
                wb.save(excelpath)
                sum_excelrow1 = sum_excelrow1 + 1
            print("%s ...Ok" %gw_name[filenum])
                
        else:
            pass    

print("\n")

# Module4 - Extract CPU Utilization #

print("### Extract CPU Utilization ###")
excelpath = "%s\summary.xlsx" %nowdir

wb = load_workbook(excelpath)
ws = wb["Summary"]

sum_excelrow1 = 8
sum_excelrow2 = 4

loglist = os.listdir(os.getcwd())

for filenum in range(0,8):
    txtfile = open("%s" %loglist[filenum], "r")
    txtline = txtfile.readlines()
    txtfile.close()

    for search_line in range(len(txtline)):
        host_pattern = r'^%s#show process cpu platform' %gw_name[filenum]
        search_result = re.match(host_pattern, txtline[search_line])

        if search_result != None:

            if sum_excelrow1 <= 11:
                stat_line = txtline[search_line + 1]
                cpusplit = stat_line.split(sep = ',',maxsplit = 2)
                fsstat = cpusplit[0]
                fssplit = fsstat.split(sep = ':  ',maxsplit = 1)
                omstat = cpusplit[1]
                omsplit = omstat.split(sep = ':  ',maxsplit = 1)
                fmstat = cpusplit[2]
                fmsplit = fmstat.split(sep = ':  ',maxsplit = 1)
                ws["N%s" %sum_excelrow1] = str(fssplit[1])
                ws["O%s" %sum_excelrow1] = str(omsplit[1])
                ws["P%s" %sum_excelrow1] = str(fmsplit[1])
                wb.save(excelpath)
                sum_excelrow1 = sum_excelrow1 + 1
                print("%s ...Ok" %loglist[filenum])
            else:
                stat_line = txtline[search_line + 1]
                cpusplit = stat_line.split(sep = ',',maxsplit = 2)
                fsstat = cpusplit[0]
                fssplit = fsstat.split(sep = ':  ',maxsplit = 1)
                omstat = cpusplit[1]
                omsplit = omstat.split(sep = ':  ',maxsplit = 1)
                fmstat = cpusplit[2]
                fmsplit = fmstat.split(sep = ':  ',maxsplit = 1)
                ws["N%s" %sum_excelrow2] = str(fssplit[1])
                ws["O%s" %sum_excelrow2] = str(omsplit[1])
                ws["P%s" %sum_excelrow2] = str(fmsplit[1])
                wb.save(excelpath)
                sum_excelrow2 = sum_excelrow2 + 1
                print("%s ...Ok" %loglist[filenum])

        else:
            pass    

print("\n")

# Module5 - Hardware Status Check #

print("### Hardware Fail Check ###")
excelpath = "%s\summary.xlsx" %nowdir

wb = load_workbook(excelpath)
ws = wb["Summary"]

sum_excelrow1 = 4
check_num = 0

loglist = os.listdir(os.getcwd())

for filenum in range(0,8):
    txtfile = open("%s" %loglist[filenum], "r")
    txtline = txtfile.readlines()
    txtfile.close()

    for search_line in range(len(txtline)):
        host_pattern = r'^%s#show platform$' %gw_name[filenum]
        search_result = re.match(host_pattern, txtline[search_line])
        if search_result != None:
            for platline in range(0,15):
                resultline = search_line + platline + 5
                if platline == 10:
                    pass
                elif platline == 11:
                    pass
                else:
                    platresult = txtline[resultline]
                    if platresult[30:32] == "ok":
                        check_num = check_num + 1
                    else:
                        print("%s - %s (%s) -> %s" %(gw_name[filenum].strip(),platresult[0:6].strip(),platresult[10:30].strip(),platresult[30:50].strip()))
                        ws["R%s" %sum_excelrow1] = str(gw_name[filenum])
                        ws["S%s" %sum_excelrow1] = str(platresult[0:6].strip())
                        ws["T%s" %sum_excelrow1] = str(platresult[10:30].strip())
                        ws["U%s" %sum_excelrow1] = str(platresult[30:50].strip())
                        wb.save(excelpath)
                        sum_excelrow1 = sum_excelrow1 + 1
        else:
            pass
    
if check_num == 104:
    print("No Defect")
    for reset_excel in range(4,12):
        ws["R%s" %reset_excel] = str("")
        ws["S%s" %reset_excel] = str("")
        ws["T%s" %reset_excel] = str("")
        ws["U%s" %reset_excel] = str("")
        wb.save(excelpath)
else:
    pass

print("\n")

# Module6 - Error Log Check #

print("### Error Log Check ###")
excelpath = "%s\summary.xlsx" %nowdir
wb = load_workbook(excelpath)
loglist = os.listdir(os.getcwd())

sheetnum = 0
dj_row = 18
doj_row = 14

for filenum in range(0,8):
    result_row = 3
    start_line = 0
    stop_line = 0
    txtfile = open("%s" %loglist[filenum])
    txtline = txtfile.readlines()
    txtfile.close()
    
    if filenum <= 3:
        sheetnum = filenum + 1
        ws = wb["대덕#%s" %sheetnum]
    else:
        sheetnum = filenum - 3
        ws = wb["보라매#%s" %sheetnum]

    for search_num in range(0,2):
        if search_num == 0:
            for search_line in range(len(txtline)):
                host_pattern = r'^%s#show logging$' %gw_name[filenum]
                search_result = re.match(host_pattern, txtline[search_line])
                if search_result != None:
                    start_line = search_line + 76
                else:
                    pass
        else:
            for search_line in range(len(txtline)):
                host_pattern = r'^%s#show interface $' %gw_name[filenum]
                search_result = re.match(host_pattern, txtline[search_line])
                if search_result != None:
                    stop_line = search_line
                else:
                    pass
    for clear_excel in range(3,500):
            ws["A%s" %clear_excel] = str("")

    for search_line in range(start_line,stop_line):
        check_log = txtline[search_line]
        split_log = check_log.split(sep = ':')
        result_log = split_log[3]
        if result_log == " %CRYPTO-4-RECVD_PKT_INV_SPI":
            pass
        elif result_log == " %CRYPTO-4-IKMP_NO_SA": 
            pass
        elif result_log == " %CRYPTO-4-IKMP_BAD_MESSAGE":
            pass
        elif result_log == " %CRYPTO-6-IKMP_MODE_FAILURE":
            pass
        elif result_log == " %CRYPTO-6-IKMP_NOT_ENCRYPTED":
            pass
        elif result_log == " %HA_EM-6-LOG":
            pass
        elif result_log == " %IOSXE-3-PLATFORM":
            repeat_split = split_log[9]
            check_repeat = repeat_split.split(sep = ' ')
            if check_repeat[1] == "%IPSEC-3-REPLAY_ERROR":
                pass
            elif check_repeat[1] == "%IPSEC-3-HMAC_ERROR":
                pass
            else:
                ws["A%s" %result_row] = str(check_log)
                result_row = result_row + 1
        else:
            ws["A%s" %result_row] = str(check_log)
            result_row = result_row + 1
    total_error = result_row - 3
    ws = wb["Summary"]
    if filenum <= 3:
        ws["X%s" %dj_row] = str(total_error)
        dj_row = dj_row + 1
    else:
        ws["X%s" %doj_row] = str(total_error)
        doj_row = doj_row + 1
    print("%s == %d" %(gw_name[filenum], total_error))
    wb.save(excelpath)